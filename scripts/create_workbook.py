#!/usr/bin/env python3
"""
Canadian Tax Preparation — Workbook Template Builder
Creates a formatted, empty Excel workbook for Claude to populate with tax data.

Usage:
  python create_workbook.py --year 2025 --name "Alex" --profile "both" --output path/to/file.xlsx

Profile options (controls which sheets are created):
  "t4"            → Sheets: Source Docs, Income, Tax Summary, Tax Estimates
  "t4+invest"     → adds: Investments (RRSP/TFSA/capital gains)
  "self-employed" → Sheets: Source Docs, Income, Expenses, Expense Detail, Tax Summary, Tax Estimates
  "both"          → All sheets including Home Office, Investments, Expenses
  "retired"       → Sheets: Source Docs, Income, Investments, Tax Summary, Tax Estimates (no business/home office)
  "simple"        → Sheets: Source Docs, Income, Tax Summary only (T4 + no investments/business)

The --sheets flag overrides profile with an explicit comma-separated list:
  --sheets "1,2,4,6,7"  → only those sheet numbers
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E4057"
LIGHT_BLUE  = "BDD7EE"
YELLOW_IN   = "FFFF99"   # cells needing user input
GREY_CALC   = "F2F2F2"   # auto-calculated
GREEN_OK    = "E2EFDA"   # confirmed / clean
RED_FLAG    = "FFE0E0"   # CPA review required
ORANGE_WARN = "FFF2CC"   # CPA attention
WHITE       = "FFFFFF"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _thin():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def h(ws, row, col, text, fill=DARK_BLUE, font_color=WHITE, bold=True, size=10,
      merge_end_col=None, wrap=False, align="center"):
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    cell.fill = _fill(fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    return cell

def d(ws, row, col, text="", fill=None, bold=False, num_fmt=None,
      align="left", wrap=False, italic=False):
    """Write a data cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", bold=bold, size=10, italic=italic)
    if fill:
        cell.fill = _fill(fill)
    if num_fmt:
        cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()
    return cell

def title_row(ws, row, text, ncols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    cell.fill = _fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    return cell

def section_row(ws, row, text, ncols=7, fill=MED_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = _fill(fill)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return cell

def input_cell(ws, row, col, hint=""):
    """Yellow cell waiting for input."""
    c = d(ws, row, col, hint or "", fill=YELLOW_IN, align="left")
    return c

def calc_cell(ws, row, col, formula="", num_fmt="$#,##0.00"):
    """Grey calculated cell."""
    c = d(ws, row, col, formula, fill=GREY_CALC, align="right", bold=True)
    c.number_format = num_fmt
    return c

# ── Sheet builders ────────────────────────────────────────────────────────────

def build_source_docs(wb, year):
    ws = wb.create_sheet("1 - Source Documents")
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40

    title_row(ws, 1, f"SOURCE DOCUMENTS — {year} TAX RETURN", ncols=6)
    for col, hdr in enumerate(
        ["Document Type","File / Description","Date Range","Extracted?","Used In Sheet","Notes / Issues"], 1
    ):
        h(ws, 2, col, hdr)

    placeholders = [
        ("T4 – Employment Income", "", "", "Pending", "2 – Income", ""),
        ("T5 / T3 – Investment Slips", "", "", "Pending", "4 – Investments", ""),
        ("RRSP Contribution Receipt", "", "", "Pending", "4 – Investments", ""),
        ("Business Bank Statements", "", "", "Pending", "2 – Income / 5 – Expenses", ""),
        ("Business Credit Card Statements", "", "", "Pending", "5 – Expenses", ""),
        ("Expense Receipts (folder)", "", "", "Pending", "5b – Expense Detail", ""),
        ("Property Tax Statement", "", "", "Pending", "3 – Home Office", ""),
        ("Mortgage / HELOC Statement", "", "", "Pending", "3 – Home Office", ""),
        ("Home Insurance Policy", "", "", "Pending", "3 – Home Office", ""),
        ("Notice of Assessment (prior year)", "", "", "Pending", "4 – Investments", "RRSP room"),
    ]
    for i, row_data in enumerate(placeholders, 3):
        for col, val in enumerate(row_data, 1):
            fill = YELLOW_IN if val == "" else None
            d(ws, i, col, val, fill=fill, wrap=True)
        ws.row_dimensions[i].height = 18

    return ws


def build_income(wb, year, profile):
    ws = wb.create_sheet("2 - Income Summary")
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 10

    title_row(ws, 1, f"INCOME SUMMARY — {year}", ncols=5)
    for col, hdr in enumerate(["Line Item","Amount (CAD)","Source","Notes","CRA Line"], 1):
        h(ws, 2, col, hdr)

    r = 3
    def income_row(label, source="", note="", cra="", input_=True):
        nonlocal r
        d(ws, r, 1, label, bold=(label.isupper() or label.startswith("TOTAL") or label.startswith("NET")))
        if input_:
            input_cell(ws, r, 2, "")
        else:
            calc_cell(ws, r, 2)
        d(ws, r, 3, source)
        d(ws, r, 4, note, wrap=True)
        d(ws, r, 5, cra, align="center")
        ws.row_dimensions[r].height = 18
        r += 1

    if profile in ("t4", "both"):
        section_row(ws, r, "  EMPLOYMENT INCOME (T4)", ncols=5); r += 1
        income_row("Gross employment income", "T4 Box 14", "", "10100")
        income_row("Taxable benefits / allowances", "T4 Box 40", "If applicable", "10400")
        income_row("RSU / stock option benefit", "T4 Box 38/39", "Already in T4 Box 14 — do NOT double-count", "10100")
        income_row("CPP employer contribution (info)", "T4 Box 16", "For reference — not added to income", "—")
        income_row("EI premiums (info)", "T4 Box 18", "Credit on Schedule 1, not deduction", "—")
        income_row("Income tax withheld (info)", "T4 Box 22", "Refundable — offsets tax owing", "43700")

    if profile in ("self-employed", "both"):
        section_row(ws, r, "  SELF-EMPLOYMENT / BUSINESS INCOME", ncols=5); r += 1
        income_row("Gross business revenue (CAD invoices)", "Bank statements / invoices", "", "13500")
        income_row("Gross business revenue (USD → CAD)", "Bank / Wise statements", "Note FX rate used", "13500")
        income_row("HST / GST collected", "", "NOT income — reduce by HST remitted", "—")
        income_row("HST / GST remitted", "", "Subtract from HST collected — net = zero", "—")
        income_row("GROSS SELF-EMPLOYMENT INCOME", "", "", "13500", input_=False)

    section_row(ws, r, "  INVESTMENT & OTHER INCOME", ncols=5); r += 1
    income_row("Eligible dividends (grossed up)", "T5 Box 10 × 1.38", "Gross-up for tax credit", "12000")
    income_row("Ordinary / ineligible dividends (grossed up)", "T5 Box 11 × 1.15", "", "12000")
    income_row("Interest income", "T5 Box 13 / bank", "", "12100")
    income_row("Capital gains (50% inclusion)", "Schedule 3", "50% of realized gains", "12700")
    income_row("RRSP withdrawals", "T4RSP", "Fully taxable if not HBP/LLP repayment", "12900")
    income_row("Employment insurance (EI)", "T4E Box 14", "", "11900")
    income_row("Other income", "", "Pensions, foreign, misc.", "13000")

    section_row(ws, r, "  TOTALS", ncols=5, fill=DARK_BLUE); r += 1
    d(ws, r, 1, "TOTAL INCOME", bold=True, fill=LIGHT_BLUE)
    calc_cell(ws, r, 2)
    d(ws, r, 5, "15000", align="center"); r += 1

    return ws


def build_home_office(wb, year):
    ws = wb.create_sheet("3 - Home Office")
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 38

    title_row(ws, 1, f"HOME OFFICE DEDUCTION — {year}", ncols=6)
    r = 2

    section_row(ws, r, "  SECTION A — HOME PROFILE", ncols=6); r += 1
    for label, hint in [
        ("Total home square footage", "e.g. 2000"),
        ("Home office square footage", "e.g. 250"),
        ("Home office percentage", "=B4/B3 (auto)"),
        ("Months used as office", "Usually 12"),
    ]:
        d(ws, r, 1, label)
        input_cell(ws, r, 2, hint)
        ws.row_dimensions[r].height = 18; r += 1

    section_row(ws, r, "  SECTION B — SIMPLIFIED METHOD ($2/sq ft)", ncols=6); r += 1
    h(ws, r, 1, "Item"); h(ws, r, 2, "Value"); h(ws, r, 3, ""); h(ws, r, 4, "Deduction"); h(ws, r, 5, "CRA Line"); h(ws, r, 6, "Notes"); r += 1
    d(ws, r, 1, "Office sq ft (from above)")
    d(ws, r, 2, "=B4", fill=GREY_CALC)
    d(ws, r, 3, "×")
    d(ws, r, 4, "$2.00/sq ft flat rate"); d(ws, r, 5, "20800"); r += 1
    d(ws, r, 1, "SIMPLIFIED DEDUCTION", bold=True)
    calc_cell(ws, r, 4); d(ws, r, 5, "20800", align="center")
    d(ws, r, 6, "Max: office sq ft × $2 (no cap for self-employed)", fill=ORANGE_WARN); r += 1

    section_row(ws, r, "  SECTION C — DETAILED METHOD", ncols=6); r += 1
    for col, hdr in enumerate(["Expense","Annual Amount ($)","Home %","Deductible ($)","CRA Line","Notes"], 1):
        h(ws, r, col, hdr); r += 1
    detailed_items = [
        ("Mortgage interest (NOT principal)", "", "=B4/(B3*B5)", "", "12500", "Interest only; get from bank statement"),
        ("OR Rent (if renting)", "", "=B4/(B3*B5)", "", "12500", ""),
        ("Property tax", "", "=B4/(B3*B5)", "", "12500", ""),
        ("Home insurance", "", "=B4/(B3*B5)", "", "12500", "Annual premium"),
        ("Utilities (hydro, gas, water)", "", "=B4/(B3*B5)", "", "12500", ""),
        ("Internet (business portion)", "", "Business %", "", "20800", "⚠️ Do not double-count with business expenses"),
        ("Maintenance / repairs", "", "=B4/(B3*B5)", "", "12500", "Repairs only, not renovations"),
        ("Condo fees (if applicable)", "", "=B4/(B3*B5)", "", "12500", ""),
    ]
    for label, amt, pct, ded, cra, note in detailed_items:
        d(ws, r, 1, label)
        input_cell(ws, r, 2, amt)
        input_cell(ws, r, 3, pct)
        calc_cell(ws, r, 4)
        d(ws, r, 5, cra, align="center")
        d(ws, r, 6, note, wrap=True)
        ws.row_dimensions[r].height = 18; r += 1
    d(ws, r, 1, "TOTAL DETAILED DEDUCTION", bold=True)
    calc_cell(ws, r, 4); d(ws, r, 5, "20800", align="center"); r += 1

    section_row(ws, r, "  SECTION D — METHOD COMPARISON", ncols=6); r += 1
    for col, hdr in enumerate(["Method","Deduction Amount","Recommended?"], 1):
        h(ws, r, col, hdr); r += 1
    d(ws, r, 1, "Simplified ($2/sq ft)")
    calc_cell(ws, r, 2); d(ws, r, 3, "Use if higher"); r += 1
    d(ws, r, 1, "Detailed method")
    calc_cell(ws, r, 2); d(ws, r, 3, "Use if higher"); r += 1
    d(ws, r, 1, "RECOMMENDED DEDUCTION", bold=True, fill=LIGHT_BLUE)
    calc_cell(ws, r, 2)
    d(ws, r, 3, "Use higher of the two above", fill=LIGHT_BLUE)

    return ws


def build_investments(wb, year):
    ws = wb.create_sheet("4 - Investments")
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 35

    title_row(ws, 1, f"INVESTMENTS & REGISTERED ACCOUNTS — {year}", ncols=7)
    r = 2

    for section, cols, rows in [
        ("  SECTION A — RRSP CONTRIBUTIONS",
         ["Description","Institution","Contribution (CAD)","Account Type","Slip?","CRA Line","Notes"],
         [("RRSP contribution (Jan–Dec)", "", "", "Own RRSP", "Yes", "20700", ""),
          ("RRSP contribution (First 60 days next year)", "", "", "Own RRSP", "Yes", "20700", "Eligible for this year's deduction"),
          ("Spousal RRSP contribution", "", "", "Spousal RRSP", "Yes", "20700", "Claimed by contributor"),
          ("RRSP room available (from NOA)", "", "", "", "", "—", "Confirm vs. total contributions"),
          ("TOTAL RRSP CONTRIBUTIONS", "", "=SUM(C3:C5)", "", "", "20700", "")]),
        ("  SECTION B — TFSA",
         ["Description","Institution","Contributions","Withdrawals","Year-End Balance","Room Used","Notes"],
         [("TFSA contributions in year", "", "", "", "", "", ""),
          ("TFSA withdrawals in year", "", "", "", "", "", "Add back as new room next year"),
          ("TFSA year-end balance", "", "", "", "", "", "No tax impact — for reference")]),
        ("  SECTION C — RESP",
         ["Description","Institution","Contributions","CESG Received","Withdrawals","Year-End Balance","Notes"],
         [("RESP contributions", "", "", "", "", "", ""),
          ("CESG grant received", "", "", "", "", "", "20% of first $2,500 = max $500/yr"),
          ("Lifetime CESG total (running)", "", "", "", "", "", "Max $7,200 lifetime")]),
    ]:
        section_row(ws, r, section, ncols=7); r += 1
        for col, hdr in enumerate(cols, 1): h(ws, r, col, hdr); r += 1
        for row_data in rows:
            for col, val in enumerate(row_data, 1):
                fill = YELLOW_IN if val == "" else (GREY_CALC if "=SUM" in str(val) else None)
                d(ws, r, col, val, fill=fill, wrap=True)
            ws.row_dimensions[r].height = 18; r += 1
        r += 1  # blank spacer

    section_row(ws, r, "  SECTION D — CAPITAL GAINS / LOSSES (Non-Registered)", ncols=7); r += 1
    note_cell = ws.cell(row=r, column=1, value="⚠️  50% INCLUSION RATE: Only 50% of net capital gains are taxable. CPA to confirm rate for 2025 (budget proposal pending).")
    note_cell.font = Font(name="Arial", italic=True, size=9, color="C00000")
    note_cell.fill = _fill(RED_FLAG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 22; r += 1
    for col, hdr in enumerate(["Security/Transaction","Account","Proceeds (CAD)","ACB/Cost (CAD)","Gain/(Loss)","50% Inclusion","Source/Slip"], 1):
        h(ws, r, col, hdr); r += 1
    for _ in range(10):
        for col in range(1, 8):
            input_cell(ws, r, col, "")
        ws.row_dimensions[r].height = 18; r += 1
    d(ws, r, 1, "NET CAPITAL GAIN / (LOSS)", bold=True, fill=LIGHT_BLUE)
    calc_cell(ws, r, 5); calc_cell(ws, r, 6)
    d(ws, r, 7, "Report 50% amount on CRA Line 12700", fill=LIGHT_BLUE); r += 1

    return ws


def build_expenses(wb, year):
    ws = wb.create_sheet("5 - Expenses")
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 42

    title_row(ws, 1, f"DEDUCTIONS & BUSINESS EXPENSES — {year}", ncols=7)
    note = ws.cell(row=2, column=1, value="All expenses below are for self-employment (T2125). Expenses reduce self-employment income. CPA has final say on deductibility.")
    note.font = Font(name="Arial", italic=True, size=9)
    note.fill = _fill(ORANGE_WARN)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 18

    for col, hdr in enumerate(["Expense Category","Description / Merchant","Amount (CAD)","50% Limit?","Receipt?","CRA Line","Notes / Source"], 1):
        h(ws, 3, col, hdr)

    r = 4
    section_row(ws, r, "  SECTION A — BUSINESS OPERATING EXPENSES (T2125)", ncols=7); r += 1

    expense_rows = [
        ("Fuel & Vehicle", "Fuel receipts + vehicle running costs", "", "No", "", "20800", "⚠️ Requires mileage log. CPA to confirm business-use %."),
        ("Vehicle Maintenance", "Car repairs, service, tires", "", "No", "", "20800", "⚠️ CPA to confirm business-use %."),
        ("Vehicle – Accessories", "Rims, accessories", "", "No", "", "20800", "⚠️ CPA REVIEW — may not qualify as deductible."),
        ("Vehicle – Car Wash", "Car washing receipts", "", "No", "", "20800", "⚠️ CPA to confirm business-use %."),
        ("Meals & Entertainment (gross)", "Client meals at restaurants", "", "Yes – 50%", "", "20800", "50% limit. Document business purpose per receipt."),
        ("Meals – 50% Deductible Amount", "50% of meals claimed above", "", "—", "", "20800", "THIS is the amount to claim on T2125, not the gross."),
        ("Travel – Airfare", "Business flights", "", "No", "", "20800", "Document business purpose."),
        ("Travel – Accommodation", "Hotels, AirBnB for business travel", "", "No", "", "20800", ""),
        ("Software & Subscriptions", "SaaS tools, AI, apps, licenses", "", "No", "", "20800", "100% if business-use only."),
        ("Professional Fees", "Accountant, lawyer, consultants", "", "No", "", "20800", ""),
        ("Professional Development", "Courses, certifications, conferences", "", "No", "", "20800", "Must be related to current profession."),
        ("Telecommunications", "Business internet, phone (business %)", "", "No", "", "20800", "⚠️ Do not double-count with home office deduction."),
        ("Office Supplies & Equipment", "Stationery, hardware, peripherals", "", "No", "", "20800", "Items > $500 may need to be capitalized (CCA)."),
        ("Business Insurance", "E&O, liability insurance", "", "No", "", "20800", ""),
        ("Banking & FX Fees", "Bank fees, wire fees, Wise FX fees", "", "No", "", "20800", ""),
        ("Gift Cards / Incentives", "Gift cards purchased", "", "No", "", "N/A", "⚠️ Generally NOT deductible per CRA. CPA REVIEW."),
        ("Insurance (Home/Auto)", "Unica and other policies", "", "No", "", "20800", "⚠️ Only business-use portion deductible. CPA to confirm."),
        ("Home Office", "See Sheet 3 — home office calculation", "", "No", "", "20800", "Link to Sheet 3 total (detailed or simplified method)."),
        ("Other Expenses", "", "", "No", "", "20800", ""),
    ]
    for row_data in expense_rows:
        for col, val in enumerate(row_data, 1):
            fill = RED_FLAG if "⚠️" in str(val) else (YELLOW_IN if (col == 3 and val == "") else None)
            d(ws, r, col, val, fill=fill, wrap=True)
        ws.row_dimensions[r].height = 20; r += 1

    d(ws, r, 1, "TOTAL BUSINESS EXPENSES (before CPA review)", bold=True, fill=LIGHT_BLUE)
    calc_cell(ws, r, 3); ws.row_dimensions[r].height = 20; r += 2

    section_row(ws, r, "  SECTION B — PERSONAL DEDUCTIONS", ncols=7); r += 1
    for col, hdr in enumerate(["Deduction","Details","Amount (CAD)","Claimant","Receipt?","CRA Line","Notes"], 1):
        h(ws, r, col, hdr); r += 1
    personal_rows = [
        ("RRSP Contributions", "See Sheet 4 — Section A", "", "", "Yes", "20700", "Total deductible contributions"),
        ("Childcare Expenses", "Daycare, nanny receipts", "", "Lower-income spouse", "Yes", "21400", "Claimed by lower-income spouse. T4 required for nanny."),
        ("Home Office Deduction", "See Sheet 3", "", "", "Yes", "20800", ""),
        ("Medical Expenses", "Receipts over 3% threshold", "", "", "No", "33099", "⚠️ Provide official receipts."),
        ("Charitable Donations", "Official charity receipts", "", "", "No", "34900", "First $200 at 15% credit; excess ~33%."),
        ("Professional / Union Dues", "T4 Box 44 or receipts", "", "", "", "21200", ""),
    ]
    for row_data in personal_rows:
        for col, val in enumerate(row_data, 1):
            fill = YELLOW_IN if (col == 3 and val == "") else None
            d(ws, r, col, val, fill=fill, wrap=True)
        ws.row_dimensions[r].height = 20; r += 1

    return ws


def build_expense_detail(wb, year):
    ws = wb.create_sheet("5b - Expense Detail")
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 48

    title_row(ws, 1, f"EXPENSE DETAIL — ALL RECEIPTS ({year})", ncols=7)
    note = ws.cell(row=2, column=1, value="⚠️  NO PII — Amounts, dates, and categories only. Source file names as references. CPA to confirm business vs. personal for each item.")
    note.font = Font(name="Arial", italic=True, size=9, color="C00000")
    note.fill = _fill(RED_FLAG)
    ws.merge_cells("A2:G2"); ws.row_dimensions[2].height = 18

    for col, hdr in enumerate(["#","Date","Source File","Category","Amount (CAD)","CRA Line","CPA Notes"], 1):
        h(ws, 3, col, hdr)

    for i in range(1, 75):
        r = i + 3
        for col in range(1, 8):
            fill = YELLOW_IN if col not in (1, 4, 6) else None
            input_cell(ws, r, col, "")
        ws.row_dimensions[r].height = 18

    return ws


def build_tax_summary(wb, year):
    ws = wb.create_sheet("6 - Tax Summary")
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 42

    title_row(ws, 1, f"TAX SUMMARY DASHBOARD — {year}", ncols=4)

    r = 2
    for col, hdr in enumerate(["Line Item","Amount (CAD)","CRA Line","Notes"], 1):
        h(ws, r, col, hdr); r += 1

    sections = [
        ("INCOME", [
            ("T4 Employment Income", "", "10100", "From Sheet 2"),
            ("Self-Employment Income (net)", "", "13500", "Revenue minus T2125 expenses"),
            ("Investment Income", "", "12100+", "Dividends + interest"),
            ("Taxable Capital Gains (50%)", "", "12700", "50% of net realized gains"),
            ("Other Income", "", "13000", "EI, pension, misc."),
            ("TOTAL INCOME", "", "15000", ""),
        ]),
        ("DEDUCTIONS", [
            ("RRSP Contributions", "", "20700", "From Sheet 4"),
            ("Home Office Deduction", "", "20800", "From Sheet 3 — higher of two methods"),
            ("Business Expenses (T2125)", "", "20800", "From Sheet 5 — T2125 total"),
            ("CPP on self-employment", "", "22200", "Schedule 8"),
            ("Child Care Expenses", "", "21400", "From Sheet 5"),
            ("Other Deductions", "", "23200", ""),
            ("TOTAL DEDUCTIONS", "", "", ""),
        ]),
        ("NET INCOME & CREDITS", [
            ("NET INCOME", "", "23600", "Total Income minus Total Deductions"),
            ("Basic personal amount credit", "(~$16,129)", "30000", "2025 federal estimate"),
            ("Dividend tax credit", "", "40425", ""),
            ("Medical expense credit", "", "33099", ""),
            ("Donation credit", "", "34900", ""),
        ]),
    ]

    for section_name, rows in sections:
        section_row(ws, r, f"  {section_name}", ncols=4); r += 1
        for label, amt, cra, note in rows:
            is_total = label.startswith("TOTAL") or label.startswith("NET")
            fill = LIGHT_BLUE if is_total else None
            d(ws, r, 1, label, bold=is_total, fill=fill)
            if is_total:
                calc_cell(ws, r, 2)
            else:
                c = d(ws, r, 2, amt, fill=YELLOW_IN if amt == "" else None, align="right")
                c.number_format = "$#,##0.00"
            d(ws, r, 3, cra, align="center")
            d(ws, r, 4, note, wrap=True)
            ws.row_dimensions[r].height = 18; r += 1
        r += 1

    # CPA questions section
    section_row(ws, r, "  KEY QUESTIONS FOR CPA", ncols=4); r += 1
    cpa_qs = [
        "1. Home office: Confirm simplified vs. detailed method and square footage.",
        "2. RRSP: Confirm total deductible contributions and available room.",
        "3. Vehicle: Confirm business-use percentage and mileage log review.",
        "4. Capital gains: Confirm ACB for transferred shares; confirm 50% inclusion rate for 2025.",
        "5. Gift cards: Are any deductible as client gifts ($300/recipient max)?",
        "6. Insurance: What portion of home/auto insurance applies to business?",
        "7. Nanny payroll: Confirm T4 filed with CRA.",
        "8. Any other questions → add here",
    ]
    for q in cpa_qs:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cell = ws.cell(row=r, column=1, value=q)
        cell.font = Font(name="Arial", size=9)
        cell.fill = _fill(ORANGE_WARN)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18; r += 1

    return ws


def build_tax_estimates(wb, year):
    ws = wb.create_sheet("7 - Tax Estimates")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 16

    title_row(ws, 1, f"ESTIMATED TAX — {year} (ONTARIO EXAMPLE)", ncols=4)
    note = ws.cell(row=2, column=1,
        value="⚠️ ESTIMATE ONLY — Not for filing. Marginal rates are approximate. CPA will calculate exact tax owing. Adjust province in notes if not Ontario.")
    note.font = Font(name="Arial", italic=True, size=9, color="C00000")
    note.fill = _fill(RED_FLAG)
    ws.merge_cells("A2:D2"); ws.row_dimensions[2].height = 24

    r = 3
    rows = [
        ("Net income (from Sheet 6)", "", "From Sheet 6, Line 23600", "$#,##0.00"),
        ("Less: RRSP deduction", "", "Reduces taxable income dollar for dollar", "$#,##0.00"),
        ("TAXABLE INCOME", "", "", "$#,##0.00"),
        ("Est. federal tax (approx.)", "", "Based on 2025 federal brackets", "$#,##0.00"),
        ("Est. Ontario provincial tax (approx.)", "", "Based on 2025 ON rates", "$#,##0.00"),
        ("Less: federal basic personal credit (~$2,419)", "", "Non-refundable credit", "$#,##0.00"),
        ("Less: ON basic personal credit (~$634)", "", "Non-refundable credit", "$#,##0.00"),
        ("Less: CPP credit", "", "From T4/schedule", "$#,##0.00"),
        ("Less: EI credit", "", "From T4", "$#,##0.00"),
        ("Less: dividend tax credit", "", "Eligible: 15.02% of grossed-up; ineligible: 9.03%", "$#,##0.00"),
        ("ESTIMATED NET TAX", "", "Rough estimate — CPA to confirm", "$#,##0.00"),
        ("Less: tax already withheld (T4 Box 22)", "", "Refundable against balance", "$#,##0.00"),
        ("ESTIMATED BALANCE OWING / (REFUND)", "", "Positive = owing; negative = refund", "$#,##0.00"),
    ]
    for label, val, note_text, fmt in rows:
        is_total = label.startswith("EST") or label.startswith("TAXABLE")
        fill = LIGHT_BLUE if is_total else (YELLOW_IN if val == "" else None)
        d(ws, r, 1, label, bold=is_total, fill=fill)
        c = d(ws, r, 2, val, fill=fill, align="right")
        c.number_format = fmt
        d(ws, r, 3, note_text, wrap=True)
        ws.row_dimensions[r].height = 18; r += 1

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Create Canadian Tax Prep workbook template")
    ap.add_argument("--year", default="2025", help="Tax year")
    ap.add_argument("--name", default="Taxpayer", help="First name (no surname)")
    ap.add_argument("--profile", default="both",
                    choices=["t4", "t4+invest", "self-employed", "both", "retired", "simple"],
                    help="Employment profile — controls which sheets are created")
    ap.add_argument("--sheets", default="",
                    help="Explicit comma-separated sheet numbers to create, e.g. '1,2,4,6,7' (overrides --profile)")
    ap.add_argument("--output", required=True, help="Output .xlsx path")
    args = ap.parse_args()

    # Determine which sheet numbers to build
    SHEET_MAP = {
        "simple":        [1, 2, 6, 7],
        "t4":            [1, 2, 6, 7],
        "t4+invest":     [1, 2, 4, 6, 7],
        "retired":       [1, 2, 4, 6, 7],
        "self-employed": [1, 2, 5, 6, 7],   # 5b added if receipts found
        "both":          [1, 2, 3, 4, 5, 6, 7],  # 5b always included for "both"
    }

    if args.sheets:
        include = set(int(x.strip()) for x in args.sheets.split(","))
    else:
        include = set(SHEET_MAP.get(args.profile, [1, 2, 3, 4, 5, 6, 7]))

    # Always include 5b alongside 5
    if 5 in include:
        include.add(55)  # 55 = "5b"

    wb = Workbook()
    wb.remove(wb.active)

    profile_for_income = "both" if args.profile in ("both", "self-employed") else "t4"

    if 1  in include: build_source_docs(wb, args.year)
    if 2  in include: build_income(wb, args.year, profile_for_income)
    if 3  in include: build_home_office(wb, args.year)
    if 4  in include: build_investments(wb, args.year)
    if 5  in include: build_expenses(wb, args.year)
    if 55 in include: build_expense_detail(wb, args.year)
    if 6  in include: build_tax_summary(wb, args.year)
    if 7  in include: build_tax_estimates(wb, args.year)

    wb.save(args.output)
    print(f"✅ Workbook created: {args.output}")
    print(f"   Year: {args.year} | Name: {args.name} | Profile: {args.profile}")
    print(f"   Sheets built: {wb.sheetnames}")


if __name__ == "__main__":
    main()
